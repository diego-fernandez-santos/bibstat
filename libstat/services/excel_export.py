# -*- coding: utf-8 -*-
from __future__ import division
import glob
import os
import datetime
import math
from django.core.files import File

from openpyxl import Workbook, load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from bibstat import settings

from data.principals import principal_for_library_type
from libstat.models import Survey, OpenData, Variable, Library

import logging

DATE_FORMAT = "%Y_%m_%d_%H_%M_%S"


logger = logging.getLogger(__name__)


def _cache_path(year, date_str=None):
    file_name = "public_export_{} {}.xslx".format(year, date_str if date_str else "*")
    if settings.ENVIRONMENT == "local":
        return "{}/data/public_exports/{}".format(os.getcwd(), file_name)
    else:
        return "/data/appl/public_exports/{}".format(file_name)


def _cached_workbook_exists_and_is_valid(year):
    paths = sorted(glob.glob(_cache_path(year)))
    if not paths:
        return False

    cache_date = datetime.datetime.strptime(paths[-1].split(" ")[-1].split(".")[0], DATE_FORMAT)
    latest_publication = OpenData.objects.first().date_modified

    return cache_date > latest_publication


def _cache_workbook(workbook, year):
    with open(_cache_path(year, datetime.datetime.utcnow().strftime(DATE_FORMAT)), "w") as f:
        File(f).write(save_virtual_workbook(workbook))


def public_excel_workbook(year):
    if not _cached_workbook_exists_and_is_valid(year):
        _cache_workbook(_published_open_data_as_workbook(year), year)
    return sorted(glob.glob(_cache_path(year)))[-1]


def _published_open_data_as_workbook(year):
    workbook = Workbook(encoding="utf-8")
    worksheet = workbook.active
    worksheet.title = u"Värden"

    variable_keys = list(OpenData.objects.filter(is_active=True, sample_year=year).distinct("variable_key"))
    sigels = list(OpenData.objects.filter(is_active=True, sample_year=year).distinct("sigel"))

    libraries = {}
    for sigel in sigels:
        libraries[sigel] = dict.fromkeys(variable_keys)

    for open_data in OpenData.objects.filter(is_active=True, sample_year=year).only("library_name", "variable_key",
                                                                                    "sigel", "value"):
        libraries[open_data.sigel][open_data.variable_key] = open_data.value

    header = ["Bibliotek", "Sigel", "Bibliotekstyp", "Kommunkod", "Stad", "Adress", "Postkod"]
    variable_index = {}
    for index, key in enumerate(variable_keys):
        variable_index[key] = index + 7
        header.append(key)
    worksheet.append(header)

    for sigel in libraries:
        library = Survey.objects.get(library__sigel=sigel).library
        row = [""] * len(header)
        row[0] = library.name
        row[1] = sigel if year >= 2014 else ""  # Do not show auto-generated sigels (used before 2014)
        row[2] = library.library_type
        row[3] = library.municipality_code
        row[4] = library.city
        row[5] = library.address
        row[6] = library.zip_code

        for key in variable_keys:
            row[variable_index[key]] = libraries[sigel][key]
        worksheet.append(row)

    variable_sheet = workbook.create_sheet()
    variable_sheet.title = u"Definitioner"
    for variable in Variable.objects.filter(key__in=variable_keys):
        variable_sheet.append([variable.key, variable.description])

    return workbook


def _load_surveys_and_append_worksheet_rows(surveys, worksheet):
    for survey in surveys:
        row = [
            survey.library.name,
            survey.library.sigel,
            survey.library.library_type,
            Survey.status_label(survey.status),
            survey.library.email,
            survey.library.municipality_code,
            survey.library.city,
            survey.library.address,
            survey.library.zip_code,
            principal_for_library_type[survey.library.library_type]
            if survey.library.library_type in principal_for_library_type else None,
            "Ja" if survey.can_publish() else "Nej: " + survey.reasons_for_not_able_to_publish(),
            "Ja" if survey.is_reporting_for_others() else "Nej",
            "Ja" if survey.is_reported_by_other() else "Nej",
            ",".join(survey.reported_by())
        ]
        for observation in survey.observations:
            row.append(observation.value)

        for sigel in survey.selected_libraries:
            if sigel != survey.library.sigel:
                survey = Survey.objects.filter(library__sigel=sigel, sample_year=survey.sample_year).first()
                row.append("%s (%s)" % (survey.library.name, survey.library.sigel))
                row.append(survey.library.address)
                row.append(survey.library.zip_code)

        worksheet.append(row)


def surveys_to_excel_workbook(survey_ids):
    bulk_size = 500
    bulks = int(math.ceil(len(survey_ids) / bulk_size))

    headers = [
        "Bibliotek",
        "Sigel",
        "Bibliotekstyp",
        "Status",
        "Email",
        "Kommunkod",
        "Stad",
        "Adress",
        "Postkod",
        "Huvudman",
        "Kan publiceras?",
        "Samredovisar",
        "Samredovisas",
        "Redovisas av"
    ]
    headers += [unicode(observation.variable.key) for observation in Survey.objects.get(pk=survey_ids[0]).observations if observation.variable.key]
    headers += ["Samredovisat bibliotek", "Gatuadress", "Postnummer"]

    workbook = Workbook(encoding="utf-8")
    worksheet = workbook.active
    worksheet.append(headers)

    for index in range(0, bulks):
        start_index = index * bulk_size
        stop_index = start_index + bulk_size
        if stop_index > len(survey_ids):
            stop_index = len(survey_ids)

        ids = survey_ids[start_index:stop_index]

        surveys = Survey.objects.filter(id__in=ids).order_by('library__name')

        _load_surveys_and_append_worksheet_rows(surveys, worksheet)

    return workbook